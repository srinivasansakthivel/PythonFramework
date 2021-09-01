import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ssakthivel\\Downloads\\TestCase.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Srinivasan"
print(sheet.cell(row=2, column=2).value)
Dict = {}
print(sheet.max_row)
print(sheet.max_column)

print(sheet["A2"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(2, sheet.max_column+1):
            # print(f"Cell value of {i}, {j} is : ", sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
